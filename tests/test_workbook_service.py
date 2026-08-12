from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from qinchecker.models.review import Confidence, FieldKey, FieldProposal, ReviewState, SpeciesRecord
from qinchecker.services.workbook import WorkbookBridge, WorkbookData, WorkbookService


class FakeBridge:
    def __init__(self, data: WorkbookData) -> None:
        self.data = data

    def read(self, _path: Path) -> WorkbookData:
        return self.data


class WorkbookServiceTests(TestCase):
    def setUp(self) -> None:
        self.headers = [field.value for field in FieldKey] + ["备注"]
        values = {
            "Family": "Salicaceae", "科": "杨柳科", "Species": "Populus old", "物种": "旧杨",
            "区县": "户县", "备注": "旧备注",
        }
        row = [values.get(header, "-") for header in self.headers]
        data = WorkbookData({"Sheet1": [self.headers, row]})
        self.service = WorkbookService(FakeBridge(data))
        self.records = self.service.load_records(Path("input.xlsx"), "Sheet1")
        self.record = self.records[0]

    def proposal(self, field: FieldKey, suggestion: object, state: ReviewState) -> FieldProposal:
        proposal = FieldProposal(
            worksheet_name="Sheet1",
            excel_row=2,
            species_name="Populus old",
            field=field,
            original_value=self.record.values[field.value],
            suggested_value=suggestion,
            final_value=suggestion if state is not ReviewState.PENDING_REVIEW else None,
            confidence=Confidence.HIGH if state is not ReviewState.PENDING_REVIEW else Confidence.MEDIUM,
            state=state,
            source_name="iPlant 中国植物志（修订版，FOC）",
            source_url="https://example.test/foc",
            source_excerpt="evidence",
            parser_rule="test rule",
            note="test note",
        )
        return proposal

    def test_load_records_respects_start_and_batch(self) -> None:
        self.assertEqual(len(self.records), 1)
        self.assertEqual(self.record.excel_row, 2)
        self.assertEqual(self.record.column_indexes["区县"], 6)

    def test_export_plan_writes_only_finalized_changes_and_appends_remark(self) -> None:
        auto = self.proposal(FieldKey.SPECIES_LATIN, "Populus new", ReviewState.AUTO_READY)
        manual = self.proposal(FieldKey.COUNTIES, "鄠邑区", ReviewState.MANUALLY_CONFIRMED)
        manual.note = "行政区名称更新：户县→鄠邑区"
        pending = self.proposal(FieldKey.HABITAT, "山坡", ReviewState.PENDING_REVIEW)
        plan = self.service.build_export_plan(self.records, [auto, manual, pending])

        written_columns = {change.column_index: change.value for change in plan.changes}
        self.assertEqual(written_columns[self.record.column_indexes["Species"]], "Populus new")
        self.assertEqual(written_columns[self.record.column_indexes["区县"]], "鄠邑区")
        self.assertNotIn(self.record.column_indexes["Habitat"], written_columns)
        remark = written_columns[self.record.column_indexes["备注"]]
        self.assertIn("旧备注", remark)
        self.assertIn("植物拉丁名由“Populus old”更新为现行接受名“Populus new”", remark)
        self.assertIn("行政区名称采用现行标准：“户县→鄠邑区”", remark)
        self.assertNotIn("Habitat", remark)
        self.assertTrue(next(change for change in plan.changes if change.column_index == 2).highlight_modified)
        self.assertEqual(len(plan.review_log), 3)
        self.assertEqual(plan.review_log[2][11], "否")

    def test_native_export_keeps_input_and_writes_audit_log(self) -> None:
        auto = self.proposal(FieldKey.SPECIES_LATIN, "Populus new", ReviewState.AUTO_READY)
        plan = self.service.build_export_plan(self.records, [auto])
        with TemporaryDirectory() as temp:
            input_path = Path(temp) / "input.xlsx"
            output_path = Path(temp) / "export.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(self.headers)
            sheet.append([self.record.values[header] for header in self.headers])
            workbook.save(input_path)
            workbook.close()

            WorkbookBridge().export(input_path, output_path, plan)
            original = load_workbook(input_path, data_only=False)
            exported = load_workbook(output_path, data_only=False)
            self.assertEqual(original["Sheet1"]["C2"].value, "Populus old")
            self.assertEqual(exported["Sheet1"]["C2"].value, "Populus new")
            self.assertEqual(exported["Sheet1"]["C2"].font.color.rgb, "FFFF0000")
            self.assertEqual(exported["Sheet1"].max_column, len(self.headers))
            self.assertEqual(exported["核对日志"]["A1"].value, "工作表")
            self.assertEqual(exported["核对日志"]["A2"].value, "Sheet1")
            original.close()
            exported.close()

    def test_summary_lists_pending_url_but_not_main_sheet_url_column(self) -> None:
        pending = self.proposal(FieldKey.HABITAT, "山坡", ReviewState.PENDING_REVIEW)
        plan = self.service.build_export_plan(self.records, [pending])
        summary = self.service.summary_text(Path("未对.xlsx"), plan)
        self.assertIn("https://example.test/foc", summary)
        self.assertIn("主表未新增来源网址列", summary)
        self.assertIn("每个植物的数据说明：", summary)
        self.assertIn("Populus old（旧杨）", summary)
        self.assertIn("已变化 0 个字段（无）", summary)
        self.assertIn("待复核 17 个字段", summary)

    def test_species_summary_counts_changed_and_pending_fields(self) -> None:
        changed = self.proposal(FieldKey.SPECIES_LATIN, "Populus new", ReviewState.AUTO_READY)
        pending = self.proposal(FieldKey.HABITAT, None, ReviewState.PENDING_REVIEW)
        unchanged = self.proposal(FieldKey.FAMILY_LATIN, "Salicaceae", ReviewState.NO_CHANGE)
        remaining = [
            self.proposal(field, self.record.values[field.value], ReviewState.NO_CHANGE)
            for field in FieldKey
            if field not in {FieldKey.SPECIES_LATIN, FieldKey.HABITAT, FieldKey.FAMILY_LATIN}
        ]
        plan = self.service.build_export_plan(self.records, [changed, pending, unchanged, *remaining])
        summary = self.service.summary_text(Path("未对.xlsx"), plan)
        self.assertIn("已变化 1 个字段（Species）", summary)
        self.assertIn("待复核 1 个字段（Habitat）", summary)

    def test_species_summary_marks_manual_import_source(self) -> None:
        proposal = self.proposal(FieldKey.HABITAT, "山坡", ReviewState.PENDING_REVIEW)
        proposal.source_name = "人工导入 · FOC文本 · 论文摘录"
        plan = self.service.build_export_plan(self.records, [proposal])
        summary = self.service.summary_text(Path("未对.xlsx"), plan)
        self.assertIn("人工导入来源：人工导入 · FOC文本 · 论文摘录", summary)

    def test_normal_review_does_not_write_a_remark(self) -> None:
        no_change = self.proposal(FieldKey.FAMILY_LATIN, "Salicaceae", ReviewState.NO_CHANGE)
        pending = self.proposal(FieldKey.HABITAT, None, ReviewState.PENDING_REVIEW)
        plan = self.service.build_export_plan(self.records, [no_change, pending])
        remark_changes = [
            change for change in plan.changes
            if change.column_index == self.record.column_indexes["备注"]
        ]
        self.assertEqual(remark_changes, [])

    def test_chinese_fallback_is_not_written_into_species_remark(self) -> None:
        proposal = self.proposal(FieldKey.SPECIES_CHINESE, "旧杨", ReviewState.NO_CHANGE)
        proposal.note = "拉丁名未搜索到有效 FOC，已改用中文名“旧杨”搜索"
        plan = self.service.build_export_plan(self.records, [proposal])
        self.assertFalse(any(
            change.column_index == self.record.column_indexes["备注"] for change in plan.changes
        ))

    def test_old_generated_review_note_is_removed_without_touching_manual_note(self) -> None:
        self.record.values["备注"] = "人工说明；QinChecker核对：旧的复核详情"
        plan = self.service.build_export_plan(self.records, [])
        remark = next(change.value for change in plan.changes if change.column_index == self.record.column_indexes["备注"])
        self.assertEqual(remark, "人工说明")
