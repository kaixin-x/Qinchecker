"""生成一份小型导出文件，验证修改色与精简备注；仅供开发回归。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from qinchecker.models.review import Confidence, FieldKey, FieldProposal, ReviewState
from qinchecker.services.workbook import WorkbookBridge, WorkbookService


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "sessions" / "style_verify"
SOURCE = OUTPUT / "样式验证_原表.xlsx"
EXPORTED = OUTPUT / "样式验证_导出.xlsx"


def proposal(record, field: FieldKey, value: object, note: str = "") -> FieldProposal:
    return FieldProposal(
        worksheet_name=record.worksheet_name,
        excel_row=record.excel_row,
        species_name=record.latin_name,
        field=field,
        original_value=record.values[field.value],
        suggested_value=value,
        final_value=value,
        confidence=Confidence.HIGH,
        state=ReviewState.AUTO_READY,
        source_url="https://www.iplant.cn/",
        source_excerpt="FOC verification text",
        parser_rule="样式验证",
        note=note,
    )


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    headers = [field.value for field in FieldKey] + ["备注"]
    values = {header: "-" for header in headers}
    values.update({"Family": "OldFamily", "Species": "Populus old", "物种": "旧杨", "区县": "户县", "备注": "人工备注"})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    sheet.append([values[header] for header in headers])
    workbook.save(SOURCE)
    workbook.close()

    service = WorkbookService(WorkbookBridge())
    record = service.load_records(SOURCE, "Sheet1", 2, 1)[0]
    latin = proposal(record, FieldKey.SPECIES_LATIN, "Populus new")
    county = proposal(record, FieldKey.COUNTIES, "鄠邑区", "行政区名称更新：户县→鄠邑区")
    kept = proposal(record, FieldKey.FAMILY_LATIN, "OldFamily")
    kept.state = ReviewState.KEPT_ORIGINAL
    kept.final_value = kept.original_value
    plan = service.build_export_plan([record], [latin, county, kept])
    service.bridge.export(SOURCE, EXPORTED, plan)

    checked = load_workbook(EXPORTED)
    result = checked["Sheet1"]
    assert result.cell(2, record.column_indexes["Species"] + 1).font.color.rgb == "FFFF0000"
    assert result.cell(2, record.column_indexes["区县"] + 1).font.color.rgb == "FFFF0000"
    assert result.cell(2, record.column_indexes["Family"] + 1).font.color.type != "rgb"
    remark = str(result.cell(2, record.column_indexes["备注"] + 1).value)
    assert "植物拉丁名" in remark and "户县→鄠邑区" in remark
    assert "已检查" not in remark and "待复核" not in remark
    checked.close()
    print(EXPORTED)


if __name__ == "__main__":
    main()
