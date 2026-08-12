"""安全读取原始 xlsx，并导出带审计日志的新工作簿。"""

from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException

from qinchecker.models.review import FieldKey, FieldProposal, ReviewState, SpeciesRecord


REMARKS_HEADER = "备注"
REVIEW_LOG_SHEET = "核对日志"
REVIEW_LOG_HEADERS = [
    "工作表", "Excel行", "物种（拉丁）", "物种（中文）", "字段", "原始值", "来源建议", "最终值",
    "复核状态", "处理动作", "置信度", "是否写入主表", "解析规则", "变更备注", "来源名称", "来源网址", "来源摘录",
]
WRITE_STATES = {
    ReviewState.AUTO_READY,
    ReviewState.ACCEPTED_SOURCE,
    ReviewState.MANUALLY_CONFIRMED,
}


class WorkbookError(RuntimeError):
    """输入工作簿无法安全读取，或导出不符合约束。"""


@dataclass(slots=True)
class WorkbookData:
    sheets: dict[str, list[list[object]]]

    @property
    def sheet_names(self) -> list[str]:
        return list(self.sheets)


@dataclass(slots=True)
class CellChange:
    worksheet_name: str
    excel_row: int
    column_index: int
    value: object
    highlight_modified: bool = False


@dataclass(frozen=True, slots=True)
class SpeciesReviewSummary:
    worksheet_name: str
    excel_row: int
    latin_name: str
    chinese_name: str
    changed_fields: tuple[str, ...]
    pending_fields: tuple[str, ...]
    source_summary: str = "iPlant网页来源"


@dataclass(slots=True)
class ExportPlan:
    changes: list[CellChange] = field(default_factory=list)
    review_log: list[list[object]] = field(default_factory=list)
    changed_proposals: list[FieldProposal] = field(default_factory=list)
    pending_proposals: list[FieldProposal] = field(default_factory=list)
    species_summaries: list[SpeciesReviewSummary] = field(default_factory=list)


class WorkbookBridge:
    """公开 Excel 引擎的薄封装；不依赖 Node、Codex 或私有运行时。"""

    def read(self, input_path: Path) -> WorkbookData:
        workbook = self._load(input_path, read_only=True)
        try:
            return WorkbookData({
                sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
                for sheet in workbook.worksheets
            })
        finally:
            workbook.close()

    def export(self, input_path: Path, output_path: Path, plan: ExportPlan) -> None:
        if input_path.resolve() == output_path.resolve():
            raise WorkbookError("为保护原始 Excel，导出文件不能与输入文件相同")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = self._load(input_path)
        try:
            for change in plan.changes:
                if change.worksheet_name not in workbook.sheetnames:
                    raise WorkbookError(f"找不到待写入工作表：{change.worksheet_name}")
                cell = workbook[change.worksheet_name].cell(
                    row=change.excel_row,
                    column=change.column_index + 1,
                    value=change.value,
                )
                if change.highlight_modified:
                    font = copy(cell.font)
                    font.color = "FFFF0000"
                    cell.font = font
            self._write_review_log(workbook, plan.review_log)
            workbook.save(output_path)
        except OSError as exc:
            raise WorkbookError(f"无法保存导出文件：{output_path}") from exc
        finally:
            workbook.close()

    @staticmethod
    def _load(input_path: Path, *, read_only: bool = False):
        if not input_path.exists():
            raise WorkbookError(f"找不到 Excel 文件：{input_path}")
        try:
            return load_workbook(input_path, read_only=read_only, data_only=False)
        except (OSError, InvalidFileException, KeyError, ValueError) as exc:
            raise WorkbookError(f"无法读取 Excel 文件：{input_path.name}") from exc

    @staticmethod
    def _write_review_log(workbook: object, review_log: list[list[object]]) -> None:
        # ``openpyxl`` 的类型没有完整暴露在旧版本类型提示中，运行期对象均为 Workbook。
        if REVIEW_LOG_SHEET in workbook.sheetnames:  # type: ignore[attr-defined]
            workbook.remove(workbook[REVIEW_LOG_SHEET])  # type: ignore[index,attr-defined]
        log_sheet = workbook.create_sheet(REVIEW_LOG_SHEET)  # type: ignore[attr-defined]
        for row in [REVIEW_LOG_HEADERS, *review_log]:
            log_sheet.append(row)

        header_fill = PatternFill("solid", fgColor="183B56")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D9E2EC")
        for cell in log_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in log_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        log_sheet.freeze_panes = "A2"
        log_sheet.sheet_view.showGridLines = False
        for column, width in {5: 28, 6: 28, 14: 34, 16: 44, 17: 60}.items():
            log_sheet.column_dimensions[chr(64 + column)].width = width
        for column in range(1, len(REVIEW_LOG_HEADERS) + 1):
            letter = chr(64 + column)
            if log_sheet.column_dimensions[letter].width is None:
                log_sheet.column_dimensions[letter].width = 16


class WorkbookService:
    """连接工作簿行、字段级建议与导出规则。"""

    def __init__(self, bridge: WorkbookBridge) -> None:
        self.bridge = bridge

    def load_records(
        self,
        input_path: Path,
        sheet_name: str,
        start_excel_row: int = 2,
        batch_size: int | None = None,
    ) -> list[SpeciesRecord]:
        data = self.bridge.read(input_path)
        rows = data.sheets.get(sheet_name)
        if not rows:
            raise WorkbookError(f"找不到工作表：{sheet_name}")
        headers = self._headers(rows, sheet_name)
        if start_excel_row < 2:
            raise WorkbookError("起始行必须不小于第 2 行")
        end_index = len(rows) if batch_size is None else min(len(rows), start_excel_row - 1 + batch_size)
        records: list[SpeciesRecord] = []
        for row_index in range(start_excel_row - 1, end_index):
            row = rows[row_index] if row_index < len(rows) else []
            values = {header: row[column] if column < len(row) else None for header, column in headers.items()}
            if not str(values.get(FieldKey.SPECIES_LATIN.value) or "").strip():
                continue
            records.append(SpeciesRecord(sheet_name, row_index + 1, values, column_indexes=headers))
        return records

    def build_export_plan(self, records: Iterable[SpeciesRecord], proposals: Iterable[FieldProposal]) -> ExportPlan:
        records_by_row = {(record.worksheet_name, record.excel_row): record for record in records}
        plan = ExportPlan()
        approved_by_row: dict[tuple[str, int], list[FieldProposal]] = defaultdict(list)
        proposals_by_row: dict[tuple[str, int], list[FieldProposal]] = defaultdict(list)
        for proposal in proposals:
            key = (proposal.worksheet_name, proposal.excel_row)
            if key not in records_by_row:
                continue
            proposals_by_row[key].append(proposal)
            plan.review_log.append(self._log_row(proposal, records_by_row[key], proposal.state in WRITE_STATES and proposal.has_change))
            if proposal.state in WRITE_STATES and proposal.has_change:
                approved_by_row[key].append(proposal)
                plan.changed_proposals.append(proposal)
            elif proposal.state in {ReviewState.PENDING_REVIEW, ReviewState.FAILED}:
                plan.pending_proposals.append(proposal)

        for key, row_proposals in approved_by_row.items():
            record = records_by_row[key]
            header_index = self._headers_from_record(record)
            for proposal in row_proposals:
                plan.changes.append(
                    CellChange(
                        record.worksheet_name,
                        record.excel_row,
                        header_index[proposal.field.value],
                        proposal.final_value,
                        highlight_modified=True,
                    )
                )

        # 主表备注只记录名称／行政区标准化；普通复核过程保留在核对日志中。
        for key, record in records_by_row.items():
            header_index = self._headers_from_record(record)
            remarks_index = header_index.get(REMARKS_HEADER)
            if remarks_index is not None:
                original = record.values.get(REMARKS_HEADER)
                merged_note = self._special_change_note(original, proposals_by_row.get(key, []))
                if merged_note is not None and merged_note != original:
                    plan.changes.append(CellChange(record.worksheet_name, record.excel_row, remarks_index, merged_note))

            row_proposals = proposals_by_row.get(key, [])
            changed_fields = {
                proposal.field for proposal in row_proposals
                if proposal.state in WRITE_STATES and proposal.has_change
            }
            covered_fields = {proposal.field for proposal in row_proposals}
            pending_fields = {
                proposal.field for proposal in row_proposals
                if proposal.state in {ReviewState.PENDING_REVIEW, ReviewState.FAILED}
            }
            pending_fields.update(set(FieldKey) - covered_fields)
            pending_fields.difference_update(changed_fields)
            manual_source_names = tuple(
                dict.fromkeys(
                    proposal.source_name
                    for proposal in row_proposals
                    if proposal.source_name.startswith("人工导入")
                )
            )
            plan.species_summaries.append(
                SpeciesReviewSummary(
                    worksheet_name=record.worksheet_name,
                    excel_row=record.excel_row,
                    latin_name=record.latin_name,
                    chinese_name=record.chinese_name,
                    changed_fields=tuple(field.value for field in FieldKey if field in changed_fields),
                    pending_fields=tuple(field.value for field in FieldKey if field in pending_fields),
                    source_summary=(
                        "人工导入来源：" + "、".join(manual_source_names)
                        if manual_source_names
                        else "iPlant网页来源"
                    ),
                )
            )
        return plan

    @staticmethod
    def summary_text(input_path: Path, plan: ExportPlan) -> str:
        state_counts = Counter(item.state.value for item in plan.changed_proposals + plan.pending_proposals)
        lines = [
            "QinChecker 核对导出说明",
            f"输入文件：{input_path.name}",
            f"写入主表的字段变更：{len(plan.changed_proposals)}",
            f"待复核或失败字段：{len(plan.pending_proposals)}",
            "说明：主表未新增来源网址列；来源网址和原文摘录位于“核对日志”工作表。",
            "",
            "待复核 / 失败明细：",
        ]
        for proposal in plan.pending_proposals:
            lines.append(
                f"- {proposal.worksheet_name}!第{proposal.excel_row}行 | {proposal.species_name} | "
                f"{proposal.field.value} | {proposal.state.value} | {proposal.note or proposal.parser_rule} | "
                f"{proposal.source_url}"
            )
        if not plan.pending_proposals:
            lines.append("- 无")
        if state_counts:
            lines.extend(["", "状态计数：", *(f"- {state}：{count}" for state, count in sorted(state_counts.items()))])
        lines.extend(["", "每个植物的数据说明："])
        for item in sorted(plan.species_summaries, key=lambda value: (value.worksheet_name, value.excel_row)):
            changed_detail = "、".join(item.changed_fields) or "无"
            pending_detail = "、".join(item.pending_fields) or "无"
            chinese_name = item.chinese_name or "-"
            lines.append(
                f"- {item.worksheet_name}!第{item.excel_row}行 | {item.latin_name}（{chinese_name}） | "
                f"{item.source_summary} | "
                f"已变化 {len(item.changed_fields)} 个字段（{changed_detail}） | "
                f"待复核 {len(item.pending_fields)} 个字段（{pending_detail}）"
            )
        if not plan.species_summaries:
            lines.append("- 无")
        return "\n".join(lines) + "\n"

    @staticmethod
    def _headers(rows: list[list[object]], sheet_name: str) -> dict[str, int]:
        if not rows:
            raise WorkbookError(f"工作表为空：{sheet_name}")
        headers = {str(value).strip(): index for index, value in enumerate(rows[0]) if str(value).strip()}
        required = {field.value for field in FieldKey} | {REMARKS_HEADER}
        missing = sorted(required - set(headers))
        if missing:
            raise WorkbookError(f"工作表“{sheet_name}”缺少必要列：{'、'.join(missing)}")
        return headers

    @staticmethod
    def _headers_from_record(record: SpeciesRecord) -> dict[str, int]:
        if not record.column_indexes:
            raise WorkbookError("记录缺少表头位置；请使用本程序导入的记录进行导出")
        return record.column_indexes

    @staticmethod
    def _special_change_note(original: object, proposals: list[FieldProposal]) -> object | None:
        def shown(value: object) -> str:
            return "-" if value in (None, "") else str(value)

        changes = [proposal for proposal in proposals if proposal.state in WRITE_STATES and proposal.has_change]
        parts: list[str] = []
        for proposal in changes:
            if proposal.field is FieldKey.SPECIES_LATIN:
                parts.append(
                    f"植物拉丁名由“{shown(proposal.original_value)}”更新为现行接受名“{shown(proposal.final_value)}”"
                )
            elif proposal.field is FieldKey.SPECIES_CHINESE:
                parts.append(
                    f"植物中文名由“{shown(proposal.original_value)}”更新为现行名称“{shown(proposal.final_value)}”"
                )
            elif proposal.field is FieldKey.COUNTIES:
                matches = []
                for old_name, new_name in re.findall(
                    r"行政区名称更新[：:]\s*([^→；、]+)→([^；、]+)", proposal.note
                ):
                    item = f"{old_name.strip()}→{new_name.strip()}"
                    if item not in matches:
                        matches.append(item)
                if matches:
                    parts.append("行政区名称采用现行标准：“" + "、".join(matches) + "”")

        current = "" if original in (None, "-") else str(original).strip()
        for marker in ("QinChecker说明：", "QinChecker核对：", "QinChecker更新："):
            if marker in current:
                current = current.split(marker, 1)[0].rstrip("；; ")
        if not parts:
            if original not in (None, "-") and current != str(original).strip():
                return current
            return None
        addition = "QinChecker说明：" + "；".join(parts)
        return f"{current}；{addition}" if current else addition

    @staticmethod
    def _log_row(proposal: FieldProposal, record: SpeciesRecord, written: bool) -> list[object]:
        return [
            proposal.worksheet_name, proposal.excel_row, proposal.species_name, record.chinese_name, proposal.field.value,
            proposal.original_value, proposal.suggested_value, proposal.final_value, proposal.state.value,
            proposal.action.value, proposal.confidence.value, "是" if written else "否", proposal.parser_rule,
            proposal.note, proposal.source_name, proposal.source_url, proposal.source_excerpt,
        ]
